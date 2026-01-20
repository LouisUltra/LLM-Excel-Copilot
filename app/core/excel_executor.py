"""
Excel 操作执行器模块
负责解析并执行 LLM 返回的操作指令
"""

import os
import re
import tempfile
import shutil
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border

# 使用 pandas 辅助复杂操作
import pandas as pd

# 尝试导入 xlrd 以支持 .xls 格式转换
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

from app.models import Operation, OperationPlan, OperationType
from app.config import settings

# 设置 matplotlib 后端（必须在 pyplot 导入前设置）
import matplotlib
matplotlib.use('Agg')  # 使用非交互式后端，适合服务器环境


class ExecutionError(Exception):
    """
    操作执行错误
    
    提供更详细的错误信息和解决建议
    """
    def __init__(self, message: str, suggestion: str = None, operation_type: str = None):
        self.message = message
        self.suggestion = suggestion
        self.operation_type = operation_type
        
        # 构建完整的错误消息
        full_message = f"❌ {message}"
        if operation_type:
            full_message = f"[{operation_type}] {full_message}"
        if suggestion:
            full_message += f"\n💡 建议: {suggestion}"
        
        super().__init__(full_message)
    
    def __str__(self):
        return self.args[0]


class ExcelExecutor:
    """
    Excel 操作执行器
    
    核心功能：
    1. 解析 LLM 返回的操作指令
    2. 使用 openpyxl/pandas 执行各类操作
    3. 保存处理后的文件
    4. 自动处理 .xls 格式（转换为 .xlsx）
    """
    
    def __init__(self, file_path: str | Path, enable_backup: bool = True):
        """
        初始化执行器
        
        Args:
            file_path: Excel 文件路径（支持 .xlsx 和 .xls）
            enable_backup: 是否启用自动备份（默认True）
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        # 操作日志（需要在调用其他方法之前初始化）
        self.operation_log: List[str] = []
        self.operation_history: List[Dict[str, Any]] = []  # 操作历史记录
        
        self._temp_file = None  # 临时文件路径（用于 .xls 转换）
        self._backup_file = None  # 备份文件路径
        self._original_extension = self.file_path.suffix.lower()
        self._enable_backup = enable_backup
        
        # 创建备份（如果启用）
        if self._enable_backup:
            self._create_backup()
        
        # 如果是 .xls 文件，先转换为 .xlsx
        if self._original_extension == '.xls':
            self._convert_xls_to_xlsx()
        
        # 加载工作簿
        self.workbook = openpyxl.load_workbook(self.file_path if not self._temp_file else self._temp_file)
        self.active_sheet = self.workbook.active
    
    def _create_backup(self):
        """创建文件备份"""
        import tempfile
        from datetime import datetime
        
        try:
            # 在临时目录创建备份
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"{self.file_path.stem}_backup_{timestamp}{self.file_path.suffix}"
            backup_dir = Path(tempfile.gettempdir()) / "excel_copilot_backups"
            backup_dir.mkdir(exist_ok=True)
            
            self._backup_file = backup_dir / backup_name
            shutil.copy2(self.file_path, self._backup_file)
            self._log(f"已创建备份: {self._backup_file.name}")
            
            # 清理旧备份（保留最近5个）
            self._cleanup_old_backups(backup_dir, keep=5)
            
        except Exception as e:
            self._log(f"警告: 备份创建失败 - {str(e)}")
            self._backup_file = None
    
    def _cleanup_old_backups(self, backup_dir: Path, keep: int = 5):
        """清理旧备份文件"""
        try:
            # 获取所有备份文件
            backups = sorted(
                backup_dir.glob(f"{self.file_path.stem}_backup_*{self.file_path.suffix}"),
                key=lambda p: p.stat().st_mtime,
                reverse=True
            )
            
            # 删除超过保留数量的旧备份
            for old_backup in backups[keep:]:
                old_backup.unlink()
                self._log(f"已清理旧备份: {old_backup.name}")
        except Exception:
            pass  # 清理失败不影响主流程
    
    def restore_from_backup(self):
        """从备份恢复文件"""
        if not self._backup_file or not self._backup_file.exists():
            raise ExecutionError("没有可用的备份文件")
        
        try:
            # 关闭当前工作簿
            if hasattr(self, 'workbook'):
                self.workbook.close()
            
            # 从备份恢复
            shutil.copy2(self._backup_file, self.file_path)
            
            # 重新加载
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.active_sheet = self.workbook.active
            
            self._log(f"已从备份恢复: {self._backup_file.name}")
            return True
        except Exception as e:
            raise ExecutionError(f"备份恢复失败: {str(e)}")
    
    def _convert_xls_to_xlsx(self):
        """将 .xls 文件转换为 .xlsx 格式的临时文件（保留所有列）"""
        if not XLRD_AVAILABLE:
            raise ExecutionError("xlrd 库未安装，无法处理 .xls 文件")
        
        try:
            # 使用 xlrd 直接读取（更底层，不会丢失列）
            xls_book = xlrd.open_workbook(self.file_path)
            
            # 创建临时 .xlsx 文件
            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd)
            self._temp_file = Path(temp_path)
            
            # 创建新的 .xlsx 工作簿
            new_workbook = openpyxl.Workbook()
            new_workbook.remove(new_workbook.active)  # 移除默认工作表
            
            # 逐个工作表复制数据
            for sheet_idx in range(xls_book.nsheets):
                xls_sheet = xls_book.sheet_by_index(sheet_idx)
                new_sheet = new_workbook.create_sheet(title=xls_sheet.name)
                
                # 复制所有行和列（包括空列）
                for row_idx in range(xls_sheet.nrows):
                    for col_idx in range(xls_sheet.ncols):
                        cell_value = xls_sheet.cell_value(row_idx, col_idx)
                        # 处理不同类型的单元格值
                        if xls_sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                            # 日期类型
                            date_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                            cell_value = datetime(*date_tuple)
                        elif xls_sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_EMPTY:
                            cell_value = None
                        elif xls_sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_BOOLEAN:
                            cell_value = bool(cell_value)
                        elif isinstance(cell_value, str):
                            # 规范化表头：去除换行符（第一行是表头）
                            if row_idx == 0:
                                cell_value = cell_value.replace('\n', '').replace('\r', '')
                        
                        new_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
            
            # 保存转换后的文件
            new_workbook.save(self._temp_file)
            self._log(f"已将 .xls 文件转换为 .xlsx 格式进行处理（保留所有列）")
            
        except Exception as e:
            raise ExecutionError(f"转换 .xls 文件失败: {str(e)}")
    
    def _log(self, message: str):
        """记录操作日志（同时打印到控制台）"""
        self.operation_log.append(message)
        print(f"    {message}")  # 实时输出到控制台
    
    def execute_plan(self, plan: OperationPlan, output_path: Optional[str | Path] = None) -> str:
        """
        执行操作计划
        
        Args:
            plan: 操作计划
            output_path: 输出文件路径，不提供则自动生成
            
        Returns:
            str: 输出文件路径
        """
        self.operation_log = []
        
        for i, operation in enumerate(plan.operations, 1):
            try:
                self._log(f"[{i}/{len(plan.operations)}] 执行: {operation.description or operation.type}")
                self._execute_operation(operation)
                self._log(f"  ✓ 完成")
            except Exception as e:
                self._log(f"  ✗ 失败: {str(e)}")
                raise ExecutionError(f"执行操作 {i} 失败: {str(e)}")
        
        # 保存文件
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"{self.file_path.stem}_processed_{timestamp}.xlsx"
            output_path = settings.output_dir / output_name
        
        output_path = Path(output_path)
        self.workbook.save(output_path)
        self._log(f"文件已保存: {output_path}")
        
        return str(output_path)
    
    def _execute_operation(self, operation: Operation):
        """执行单个操作"""
        # 记录操作历史
        operation_record = {
            "type": operation.type.value,
            "description": operation.description,
            "timestamp": datetime.now().isoformat(),
            "target_sheet": operation.target_sheet or "active"
        }
        
        try:
            # 获取目标工作表
            if operation.target_sheet and operation.target_sheet in self.workbook.sheetnames:
                sheet = self.workbook[operation.target_sheet]
            else:
                sheet = self.active_sheet
            
            # 根据操作类型分发
            executor_map = {
                OperationType.FILTER: self._execute_filter,
                OperationType.SORT: self._execute_sort,
                OperationType.ADD_COLUMN: self._execute_add_column,
                OperationType.DELETE_COLUMN: self._execute_delete_column,
                OperationType.DELETE_ROWS: self._execute_delete_rows,
                OperationType.DEDUPLICATE: self._execute_deduplicate,
                OperationType.CALCULATE: self._execute_calculate,
                OperationType.FORMAT: self._execute_format,
                OperationType.STYLE: self._execute_style,
                OperationType.REPLACE: self._execute_replace,
                OperationType.FILL: self._execute_fill,
                OperationType.SPLIT_COLUMN: self._execute_split_column,
                OperationType.MERGE_COLUMNS: self._execute_merge_columns,
                OperationType.VLOOKUP: self._execute_vlookup,
                OperationType.PIVOT: self._execute_pivot,
                OperationType.CREATE_CHART: self._execute_create_chart,
                OperationType.MERGE_VERTICAL: self._execute_merge_vertical,
                OperationType.MERGE_HORIZONTAL: self._execute_merge_horizontal,
            }
            
            executor = executor_map.get(operation.type)
            if executor:
                executor(sheet, operation.params)
                operation_record["status"] = "success"
            else:
                raise ExecutionError(f"不支持的操作类型: {operation.type}")
                
        except Exception as e:
            operation_record["status"] = "failed"
            operation_record["error"] = str(e)
            raise
        finally:
            self.operation_history.append(operation_record)
    
    def _get_column_index(self, sheet: Worksheet, column_name: str) -> int:
        """根据列名获取列索引(1-based)，提供友好的错误提示"""
        import re
        # 规范化表头：去除换行符
        headers = [str(cell.value).replace('\n', '').replace('\r', '') if cell.value else '' for cell in sheet[1]]
        # 也规范化要查找的列名
        normalized_column_name = str(column_name).replace('\n', '').replace('\r', '')
        
        # 1. 尝试直接匹配
        if normalized_column_name in headers:
            return headers.index(normalized_column_name) + 1
        
        # 2. 尝试格式转换匹配（.N -> _N 或 _N -> .N）
        alt_name = re.sub(r'\.(\d+)$', r'_\1', normalized_column_name)  # .1 -> _1
        if alt_name != normalized_column_name and alt_name in headers:
            self._log(f"  自动修正列名: '{normalized_column_name}' -> '{alt_name}'")
            return headers.index(alt_name) + 1
        
        alt_name = re.sub(r'_(\d+)$', r'.\1', normalized_column_name)  # _1 -> .1
        if alt_name != normalized_column_name and alt_name in headers:
            self._log(f"  自动修正列名: '{normalized_column_name}' -> '{alt_name}'")
            return headers.index(alt_name) + 1
        
        # 3. 尝试大小写不敏感匹配
        lower_headers = [h.lower() for h in headers]
        if normalized_column_name.lower() in lower_headers:
            actual_name = headers[lower_headers.index(normalized_column_name.lower())]
            self._log(f"  自动修正列名大小写: '{normalized_column_name}' -> '{actual_name}'")
            return lower_headers.index(normalized_column_name.lower()) + 1
        
        # 4. 尝试模糊匹配（包含关系）
        for i, h in enumerate(headers):
            if h and (normalized_column_name.lower() in h.lower() or h.lower() in normalized_column_name.lower()):
                self._log(f"  模糊匹配列名: '{normalized_column_name}' -> '{h}'")
                return i + 1
        
        # 构建友好的错误提示
        similar_columns = [h for h in headers if h and (
            normalized_column_name.lower() in str(h).lower() or 
            str(h).lower() in normalized_column_name.lower()
        )]
        
        available_cols = [h for h in headers if h][:10]
        
        if similar_columns:
            suggestion = f"检查列名拼写，或使用相似的列名：{', '.join(similar_columns[:3])}"
        else:
            suggestion = f"请从可用列中选择：{', '.join(available_cols)}"
        
        raise ExecutionError(
            f"找不到列 '{normalized_column_name}'",
            suggestion=suggestion,
            operation_type="列名验证"
        )
    
    def _sheet_to_dataframe(self, sheet: Worksheet) -> pd.DataFrame:
        """将工作表转换为 DataFrame，处理重复列名和换行符"""
        data = list(sheet.values)
        if not data:
            return pd.DataFrame()
        
        # 规范化列名：去除换行符
        columns = [str(c).replace('\n', '').replace('\r', '') if c else '' for c in data[0]]
        
        # 处理重复的列名：给重复的列名添加后缀
        seen = {}
        unique_columns = []
        for col in columns:
            col_str = str(col) if col is not None else "Unnamed"
            if col_str in seen:
                seen[col_str] += 1
                unique_columns.append(f"{col_str}_{seen[col_str]}")
            else:
                seen[col_str] = 0
                unique_columns.append(col_str)
        
        return pd.DataFrame(data[1:], columns=unique_columns)
    
    def _sheet_to_dataframe_with_values(self, sheet: Worksheet) -> pd.DataFrame:
        """将工作表转换为 DataFrame，自动计算公式值"""
        # 获取列名
        headers = [cell.value for cell in sheet[1]]
        columns = [str(c).replace('\n', '').replace('\r', '') if c else f'Col{i}' for i, c in enumerate(headers)]
        
        # 处理重复列名
        seen = {}
        unique_columns = []
        for col in columns:
            if col in seen:
                seen[col] += 1
                unique_columns.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                unique_columns.append(col)
        
        # 提取数据，评估公式
        data_rows = []
        for row_idx in range(2, sheet.max_row + 1):
            row_data = []
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # 如果是公式，尝试获取计算值或手动计算
                if isinstance(value, str) and value.startswith('='):
                    # 尝试从缓存获取计算值
                    if hasattr(cell, 'value') and cell.data_type == 'f':
                        # 有些cell可能有cached_value
                        pass
                    # 尝试手动计算简单公式
                    calculated = self._evaluate_simple_formula(sheet, value)
                    if calculated is not None:
                        value = calculated
                
                row_data.append(value)
            data_rows.append(row_data)
        
        return pd.DataFrame(data_rows, columns=unique_columns)
    
    def _evaluate_simple_formula(self, sheet: Worksheet, formula: str):
        """
        计算Excel公式
        
        支持的功能：
        - 四则运算: +、-、*、/
        - 聚合函数: SUM, AVERAGE, COUNT, MAX, MIN
        - 条件函数: IF (基础版)
        - 嵌套公式
        """
        import re
        
        if not formula.startswith('='):
            return None
        
        formula_body = formula[1:].strip()
        
        try:
            # 1. 处理简单的算术运算（如: B2+C2、B2*C2）
            if re.match(r'^[A-Z]+\d+(\s*[\+\-\*\/]\s*[A-Z]+\d+)*$', formula_body, re.IGNORECASE):
                return self._eval_arithmetic(sheet, formula_body)
            
            # 2. 处理SUM函数
            if formula_body.upper().startswith('SUM('):
                return self._eval_sum(sheet, formula_body)
            
            # 3. 处理AVERAGE函数
            if formula_body.upper().startswith('AVERAGE('):
                return self._eval_average(sheet, formula_body)
            
            # 4. 处理COUNT函数
            if formula_body.upper().startswith('COUNT('):
                return self._eval_count(sheet, formula_body)
            
            # 5. 处理MAX函数
            if formula_body.upper().startswith('MAX('):
                return self._eval_max(sheet, formula_body)
            
            # 6. 处理MIN函数
            if formula_body.upper().startswith('MIN('):
                return self._eval_min(sheet, formula_body)
            
            # 7. 处理简单的IF函数
            if formula_body.upper().startswith('IF('):
                return self._eval_if(sheet, formula_body)
            
        except Exception as e:
            # 静默失败，返回None让openpyxl处理原始公式
            pass
        
        return None
    
    def _get_cell_value(self, sheet: Worksheet, cell_ref: str):
        """获取单元格值，递归处理公式"""
        try:
            cell_value = sheet[cell_ref].value
            if isinstance(cell_value, str) and cell_value.startswith('='):
                return self._evaluate_simple_formula(sheet, cell_value)
            return cell_value
        except Exception:
            return None
    
    def _eval_arithmetic(self, sheet: Worksheet, formula_body: str):
        """评估算术表达式"""
        import re
        
        def replace_cell_ref(match):
            col_letter = match.group(1).upper()
            row_num = match.group(2)
            cell_ref = f"{col_letter}{row_num}"
            cell_value = self._get_cell_value(sheet, cell_ref)
            
            if cell_value is not None and not isinstance(cell_value, str):
                return str(float(cell_value))
            return '0'
        
        # 替换所有单元格引用为数值
        expression = re.sub(r'([A-Z]+)(\d+)', replace_cell_ref, formula_body, flags=re.IGNORECASE)
        result = eval(expression)
        return result
    
    def _parse_range(self, range_str: str):
        """解析范围字符串如 'B2:B10' -> ('B', 2, 'B', 10)"""
        import re
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_str, re.IGNORECASE)
        if match:
            return (match.group(1).upper(), int(match.group(2)), 
                   match.group(3).upper(), int(match.group(4)))
        return None
    
    def _get_range_values(self, sheet: Worksheet, range_str: str):
        """获取范围内的所有值"""
        parsed = self._parse_range(range_str)
        if not parsed:
            return []
        
        col_start, row_start, col_end, row_end = parsed
        values = []
        
        # 只支持单列或单行范围
        if col_start == col_end:
            # 单列
            for row_num in range(row_start, row_end + 1):
                val = self._get_cell_value(sheet, f"{col_start}{row_num}")
                if val is not None and not isinstance(val, str):
                    values.append(float(val))
        elif row_start == row_end:
            # 单行
            from openpyxl.utils import column_index_from_string
            col_start_idx = column_index_from_string(col_start)
            col_end_idx = column_index_from_string(col_end)
            for col_idx in range(col_start_idx, col_end_idx + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                val = self._get_cell_value(sheet, f"{col_letter}{row_start}")
                if val is not None and not isinstance(val, str):
                    values.append(float(val))
        
        return values
    
    def _eval_sum(self, sheet: Worksheet, formula_body: str):
        """评估SUM函数"""
        import re
        match = re.match(r'^SUM\(([A-Z]+\d+:[A-Z]+\d+)\)$', formula_body, re.IGNORECASE)
        if match:
            range_str = match.group(1)
            values = self._get_range_values(sheet, range_str)
            return sum(values) if values else 0
        return None
    
    def _eval_average(self, sheet: Worksheet, formula_body: str):
        """评估AVERAGE函数"""
        import re
        match = re.match(r'^AVERAGE\(([A-Z]+\d+:[A-Z]+\d+)\)$', formula_body, re.IGNORECASE)
        if match:
            range_str = match.group(1)
            values = self._get_range_values(sheet, range_str)
            return sum(values) / len(values) if values else 0
        return None
    
    def _eval_count(self, sheet: Worksheet, formula_body: str):
        """评估COUNT函数"""
        import re
        match = re.match(r'^COUNT\(([A-Z]+\d+:[A-Z]+\d+)\)$', formula_body, re.IGNORECASE)
        if match:
            range_str = match.group(1)
            values = self._get_range_values(sheet, range_str)
            return len(values)
        return None
    
    def _eval_max(self, sheet: Worksheet, formula_body: str):
        """评估MAX函数"""
        import re
        match = re.match(r'^MAX\(([A-Z]+\d+:[A-Z]+\d+)\)$', formula_body, re.IGNORECASE)
        if match:
            range_str = match.group(1)
            values = self._get_range_values(sheet, range_str)
            return max(values) if values else 0
        return None
    
    def _eval_min(self, sheet: Worksheet, formula_body: str):
        """评估MIN函数"""
        import re
        match = re.match(r'^MIN\(([A-Z]+\d+:[A-Z]+\d+)\)$', formula_body, re.IGNORECASE)
        if match:
            range_str = match.group(1)
            values = self._get_range_values(sheet, range_str)
            return min(values) if values else 0
        return None
    
    def _eval_if(self, sheet: Worksheet, formula_body: str):
        """评估简单的IF函数，如: IF(A1>10, "大", "小")"""
        import re
        # 匹配: IF(条件, 值1, 值2)
        match = re.match(r'^IF\((.+?),(.+?),(.+?)\)$', formula_body, re.IGNORECASE)
        if match:
            condition_str = match.group(1).strip()
            value_true = match.group(2).strip()
            value_false = match.group(3).strip()
            
            # 评估条件（简单的比较运算）
            # 替换单元格引用
            def replace_ref(m):
                cell_ref = f"{m.group(1).upper()}{m.group(2)}"
                val = self._get_cell_value(sheet, cell_ref)
                return str(val) if val is not None else '0'
            
            condition = re.sub(r'([A-Z]+)(\d+)', replace_ref, condition_str, flags=re.IGNORECASE)
            
            try:
                # 评估条件
                result = eval(condition)
                
                # 返回对应的值
                if result:
                    # 去掉引号
                    return value_true.strip('"\'') if value_true.startswith(('"', "'")) else float(value_true)
                else:
                    return value_false.strip('"\'') if value_false.startswith(('"', "'")) else float(value_false)
            except Exception:
                pass
        
        return None
    
    def _dataframe_to_sheet(self, df: pd.DataFrame, sheet: Worksheet):
        """将 DataFrame 写回工作表，正确处理 NaN 值"""
        import numpy as np
        
        # 计算需要删除的行数
        old_max_row = sheet.max_row
        new_row_count = len(df) + 1  # +1 for header
        
        # 先删除所有数据行再写入（更可靠的方式）
        # 从最后一行开始往前删除，直到只剩表头
        if old_max_row > new_row_count:
            for row_idx in range(old_max_row, new_row_count, -1):
                sheet.delete_rows(row_idx)
        
        # 写入新数据
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                # 处理 NaN 值和特殊类型
                if pd.isna(value):
                    cell_value = None
                elif isinstance(value, (np.integer, np.floating)):
                    # 将 numpy 类型转换为 Python 原生类型
                    cell_value = value.item() if not pd.isna(value) else None
                elif isinstance(value, np.ndarray):
                    cell_value = str(value)
                else:
                    cell_value = value
                
                sheet.cell(row=r_idx, column=c_idx, value=cell_value)
    
    # ============ 操作实现 ============
    
    def _execute_filter(self, sheet: Worksheet, params: Dict[str, Any]):
        """执行筛选操作"""
        column = params.get("column")
        operator = params.get("operator", "eq")
        value = params.get("value")
        
        # 参数验证
        if not column:
            raise ExecutionError(
                "筛选操作缺少列名参数",
                suggestion="请指定要筛选的列名，例如：column='销售额'",
                operation_type="FILTER"
            )
        if value is None:
            raise ExecutionError(
                "筛选操作缺少筛选值",
                suggestion="请指定筛选条件的值，例如：value=1000",
                operation_type="FILTER"
            )
        
        valid_operators = ["eq", "ne", "gt", "lt", "gte", "lte", "contains", "startswith", "endswith"]
        if operator not in valid_operators:
            raise ExecutionError(
                f"不支持的筛选操作符 '{operator}'",
                suggestion=f"请使用支持的操作符：{', '.join(valid_operators)}",
                operation_type="FILTER"
            )
        
        try:
            col_idx = self._get_column_index(sheet, column)
        except ExecutionError as e:
            raise ExecutionError(
                f"筛选操作列名错误: {str(e)}",
                operation_type="FILTER"
            )
        
        # 使用 pandas 进行筛选
        try:
            df = self._sheet_to_dataframe(sheet)
            
            if operator == "eq":
                mask = df[column] == value
            elif operator == "ne":
                mask = df[column] != value
            elif operator == "gt":
                mask = pd.to_numeric(df[column], errors='coerce') > float(value)
            elif operator == "lt":
                mask = pd.to_numeric(df[column], errors='coerce') < float(value)
            elif operator == "gte":
                mask = pd.to_numeric(df[column], errors='coerce') >= float(value)
            elif operator == "lte":
                mask = pd.to_numeric(df[column], errors='coerce') <= float(value)
            elif operator == "contains":
                mask = df[column].astype(str).str.contains(str(value), case=False, na=False)
            elif operator == "startswith":
                mask = df[column].astype(str).str.startswith(str(value), na=False)
            elif operator == "endswith":
                mask = df[column].astype(str).str.endswith(str(value), na=False)
            else:
                raise ExecutionError(f"不支持的操作符: {operator}")
            
            filtered_df = df[mask]
            self._dataframe_to_sheet(filtered_df, sheet)
            self._log(f"  筛选后保留 {len(filtered_df)}/{len(df)} 行")
            
        except Exception as e:
            if isinstance(e, ExecutionError):
                raise
            raise ExecutionError(f"筛选操作执行失败: {str(e)}")
    
    def _execute_sort(self, sheet: Worksheet, params: Dict[str, Any]):
        """执行排序操作"""
        column = params.get("column")
        order = params.get("order", "asc")
        
        df = self._sheet_to_dataframe(sheet)
        df = df.sort_values(by=column, ascending=(order == "asc"))
        self._dataframe_to_sheet(df, sheet)
    
    def _execute_add_column(self, sheet: Worksheet, params: Dict[str, Any]):
        """新增列"""
        name = params.get("name")
        formula = params.get("formula", "")
        position = params.get("position", "end")
        
        # #region agent log
        import json
        with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
            f.write(json.dumps({"location":"excel_executor.py:455","message":"add_column_start","data":{"name":name,"formula":formula,"position":position},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"B"}) + '\n')
        # #endregion
        
        # 确定插入位置
        if position == "end":
            col_idx = sheet.max_column + 1
        elif position.startswith("after:"):
            ref_col = position[6:]
            col_idx = self._get_column_index(sheet, ref_col) + 1
        elif position.startswith("before:"):
            ref_col = position[7:]
            col_idx = self._get_column_index(sheet, ref_col)
        else:
            col_idx = sheet.max_column + 1
        
        # #region agent log
        new_col_letter = get_column_letter(col_idx)
        with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
            f.write(json.dumps({"location":"excel_executor.py:476","message":"column_position","data":{"col_idx":col_idx,"col_letter":new_col_letter,"max_column_before":sheet.max_column},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"B"}) + '\n')
        # #endregion
        
        # 插入列
        sheet.insert_cols(col_idx)
        sheet.cell(row=1, column=col_idx, value=name)
        
        # 如果有公式，填充公式
        if formula:
            # #region agent log
            sample_formula = self._adjust_formula_row(formula, 2)
            with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
                f.write(json.dumps({"location":"excel_executor.py:491","message":"formula_generation","data":{"original_formula":formula,"sample_row2_formula":sample_formula,"target_column":new_col_letter},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"B"}) + '\n')
            # #endregion
            
            for row in range(2, sheet.max_row + 1):
                # 替换公式中的行号引用
                row_formula = self._adjust_formula_row(formula, row)
                sheet.cell(row=row, column=col_idx, value=row_formula)
    
    def _adjust_formula_row(self, formula: str, row: int) -> str:
        """调整公式中的行号引用"""
        # 将公式中的数字行号替换为当前行号
        # 例如: =A2+B2 -> =A{row}+B{row}
        def replace_row(match):
            col = match.group(1)
            return f"{col}{row}"
        
        return re.sub(r'([A-Z]+)\d+', replace_row, formula)
    
    def _execute_delete_column(self, sheet: Worksheet, params: Dict[str, Any]):
        """删除列"""
        columns = params.get("columns", [])
        if isinstance(columns, str):
            columns = [columns]
        
        # 从右到左删除，避免索引错乱
        col_indices = sorted(
            [self._get_column_index(sheet, col) for col in columns],
            reverse=True
        )
        
        # 在删除列之前，将所有公式转换为值，防止循环引用
        # （openpyxl 删除列时不会自动更新公式引用）
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    # 尝试计算公式值
                    try:
                        # 使用 _evaluate_simple_formula 计算
                        calculated = self._evaluate_simple_formula(sheet, cell.value)
                        if calculated is not None:
                            cell.value = calculated
                    except Exception:
                        pass  # 保留原公式
        
        for col_idx in col_indices:
            sheet.delete_cols(col_idx)
    
    def _execute_delete_rows(self, sheet: Worksheet, params: Dict[str, Any]):
        """删除符合条件的行"""
        condition = params.get("condition", {})
        column = condition.get("column")
        operator = condition.get("operator", "eq")
        value = condition.get("value")
        
        col_idx = self._get_column_index(sheet, column)
        
        # 从下往上删除，避免索引错乱
        rows_to_delete = []
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=col_idx).value
            if self._matches_condition(cell_value, operator, value):
                rows_to_delete.append(row)
        
        for row in reversed(rows_to_delete):
            sheet.delete_rows(row)
        
        self._log(f"  删除了 {len(rows_to_delete)} 行")
    
    def _matches_condition(self, cell_value: Any, operator: str, value: Any) -> bool:
        """检查值是否符合条件（支持宽松的布尔值匹配）"""
        if cell_value is None:
            cell_value = ""
        
        str_value = str(cell_value).strip()
        str_target = str(value).strip()
        
        # 布尔值标准化（宽松匹配）
        def normalize_bool(v):
            """将各种布尔值格式统一化"""
            v_lower = str(v).strip().lower()
            if v_lower in ['true', '1', 'yes', 'y', '是', '真']:
                return 'TRUE'
            elif v_lower in ['false', '0', 'no', 'n', '否', '假', '']:
                return 'FALSE'
            return v
        
        # 对于 eq 和 ne，先尝试布尔值标准化
        if operator in ["eq", "ne"]:
            normalized_value = normalize_bool(str_value)
            normalized_target = normalize_bool(str_target)
            
            if operator == "eq":
                return normalized_value == normalized_target
            else:  # ne
                return normalized_value != normalized_target
        
        if operator == "contains":
            return str_target.lower() in str_value.lower()
        elif operator == "empty":
            return str_value == ""
        elif operator == "not_empty":
            return str_value != ""
        
        # 数值比较
        try:
            num_cell = float(str_value.replace(",", ""))
            num_value = float(str_target)
            if operator == "gt":
                return num_cell > num_value
            elif operator == "lt":
                return num_cell < num_value
            elif operator == "gte":
                return num_cell >= num_value
            elif operator == "lte":
                return num_cell <= num_value
        except ValueError:
            pass
        
        return False
    
    def _execute_deduplicate(self, sheet: Worksheet, params: Dict[str, Any]):
        """去重"""
        columns = params.get("columns", [])
        keep = params.get("keep", "first")
        
        df = self._sheet_to_dataframe(sheet)
        original_count = len(df)
        
        if columns:
            df = df.drop_duplicates(subset=columns, keep=keep)
        else:
            df = df.drop_duplicates(keep=keep)
        
        self._dataframe_to_sheet(df, sheet)
        self._log(f"  去重删除了 {original_count - len(df)} 行")
    
    def _execute_calculate(self, sheet: Worksheet, params: Dict[str, Any]):
        """计算汇总"""
        operations = params.get("operations", [])
        
        # 在末尾添加汇总行
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value="汇总")
        
        # #region agent log
        import json
        with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
            f.write(json.dumps({"location":"excel_executor.py:617","message":"calculate_start","data":{"next_row":next_row,"operations_count":len(operations)},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"B"}) + '\n')
        # #endregion
        
        for op in operations:
            column = op.get("column")
            function = op.get("function", "sum")
            col_idx = self._get_column_index(sheet, column)
            col_letter = get_column_letter(col_idx)
            
            # 插入汇总公式
            if function == "sum":
                formula = f"=SUM({col_letter}2:{col_letter}{next_row-1})"
            elif function == "avg":
                formula = f"=AVERAGE({col_letter}2:{col_letter}{next_row-1})"
            elif function == "count":
                formula = f"=COUNT({col_letter}2:{col_letter}{next_row-1})"
            elif function == "max":
                formula = f"=MAX({col_letter}2:{col_letter}{next_row-1})"
            elif function == "min":
                formula = f"=MIN({col_letter}2:{col_letter}{next_row-1})"
            else:
                continue
            
            # #region agent log
            with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
                f.write(json.dumps({"location":"excel_executor.py:641","message":"formula_inserted","data":{"column":column,"function":function,"formula":formula,"row":next_row},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"B"}) + '\n')
            # #endregion
            
            sheet.cell(row=next_row, column=col_idx, value=formula)
    
    def _execute_format(self, sheet: Worksheet, params: Dict[str, Any]):
        """格式化（数字/日期格式）"""
        column = params.get("column")
        format_type = params.get("format_type")
        format_string = params.get("format_string", "")
        
        # 如果没有指定列，跳过（可能是样式操作被错误路由到这里）
        if not column:
            self._log("  警告: 未指定列名，跳过格式化")
            return
        
        col_idx = self._get_column_index(sheet, column)
        
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_idx)
            
            if format_type == "number":
                cell.number_format = format_string or "#,##0.00"
            elif format_type == "date":
                cell.number_format = format_string or "yyyy-mm-dd"
            elif format_type == "percentage":
                cell.number_format = format_string or "0.00%"
            elif format_type == "currency":
                cell.number_format = format_string or "¥#,##0.00"
    
    def _execute_style(self, sheet: Worksheet, params: Dict[str, Any]):
        """设置样式（边框、背景色等）"""
        from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
        
        style_type = params.get("style_type", "border")  # border, fill, header, all
        range_str = params.get("range", "")  # 例如 "A1:L228"
        header_row = params.get("header_row", 1)  # 标题行号
        border_style = params.get("border_style", "thin")  # thin, medium, thick
        fill_color = params.get("fill_color", "D9E1F2")  # 十六进制颜色
        
        # 定义边框样式
        border_styles = {
            "thin": Side(style="thin", color="000000"),
            "medium": Side(style="medium", color="000000"),
            "thick": Side(style="thick", color="000000"),
        }
        side = border_styles.get(border_style, border_styles["thin"])
        border = Border(left=side, right=side, top=side, bottom=side)
        
        # 定义填充样式
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # 确定操作范围
        if range_str:
            # 解析范围字符串，如 "A1:L228"
            try:
                min_col, min_row, max_col, max_row = self._parse_range(range_str)
            except Exception:
                # 使用整个数据区域
                min_row, min_col = 1, 1
                max_row, max_col = sheet.max_row, sheet.max_column
        else:
            # 默认使用整个数据区域
            min_row, min_col = 1, 1
            max_row, max_col = sheet.max_row, sheet.max_column
        
        cells_styled = 0
        
        if style_type in ["border", "all"]:
            # 添加边框
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.border = border
                    cells_styled += 1
            self._log(f"  已为 {cells_styled} 个单元格添加边框")
        
        if style_type in ["fill", "header", "all"]:
            # 设置标题行背景色
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=header_row, column=col)
                cell.fill = fill
                # 标题行通常需要加粗
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            self._log(f"  已为标题行设置背景色和样式")
        
        if style_type == "fill" and params.get("all_rows", False):
            # 为所有行设置背景色（较少使用）
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.fill = fill
    
    def _parse_range(self, range_str: str) -> Tuple[int, int, int, int]:
        """解析 Excel 范围字符串，返回 (min_col, min_row, max_col, max_row)"""
        import re
        from openpyxl.utils import column_index_from_string
        
        # 匹配如 "A1:L228" 格式
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_str.upper())
        if match:
            min_col = column_index_from_string(match.group(1))
            min_row = int(match.group(2))
            max_col = column_index_from_string(match.group(3))
            max_row = int(match.group(4))
            return min_col, min_row, max_col, max_row
        
        raise ValueError(f"无法解析范围: {range_str}")
    
    def _execute_replace(self, sheet: Worksheet, params: Dict[str, Any]):
        """替换"""
        column = params.get("column")
        old_value = params.get("old_value")
        new_value = params.get("new_value")
        use_regex = params.get("regex", False)
        
        col_idx = self._get_column_index(sheet, column)
        count = 0
        
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_idx)
            if cell.value is None:
                continue
            
            str_value = str(cell.value)
            if use_regex:
                new_str = re.sub(old_value, new_value, str_value)
            else:
                new_str = str_value.replace(str(old_value), str(new_value))
            
            if new_str != str_value:
                cell.value = new_str
                count += 1
        
        self._log(f"  替换了 {count} 处")
    
    def _execute_fill(self, sheet: Worksheet, params: Dict[str, Any]):
        """填充空值"""
        column = params.get("column")
        method = params.get("method", "value")
        fill_value = params.get("value", "")
        
        col_idx = self._get_column_index(sheet, column)
        
        if method == "value":
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col_idx)
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = fill_value
        elif method == "ffill":
            # 向前填充
            last_value = None
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col_idx)
                if cell.value is not None and str(cell.value).strip() != "":
                    last_value = cell.value
                elif last_value is not None:
                    cell.value = last_value
        elif method == "bfill":
            # 向后填充
            last_value = None
            for row in range(sheet.max_row, 1, -1):
                cell = sheet.cell(row=row, column=col_idx)
                if cell.value is not None and str(cell.value).strip() != "":
                    last_value = cell.value
                elif last_value is not None:
                    cell.value = last_value
    
    def _execute_split_column(self, sheet: Worksheet, params: Dict[str, Any]):
        """拆分列"""
        column = params.get("column")
        delimiter = params.get("delimiter", " ")
        new_columns = params.get("new_columns", [])
        
        col_idx = self._get_column_index(sheet, column)
        
        # 使用 pandas 处理
        df = self._sheet_to_dataframe(sheet)
        split_result = df[column].astype(str).str.split(delimiter, expand=True)
        
        # 添加新列
        for i, new_col_name in enumerate(new_columns):
            if i < split_result.shape[1]:
                df[new_col_name] = split_result[i]
        
        self._dataframe_to_sheet(df, sheet)
    
    def _execute_merge_columns(self, sheet: Worksheet, params: Dict[str, Any]):
        """合并列"""
        columns = params.get("columns", [])
        new_name = params.get("new_name", "合并列")
        delimiter = params.get("delimiter", " ")
        
        df = self._sheet_to_dataframe(sheet)
        df[new_name] = df[columns].astype(str).agg(delimiter.join, axis=1)
        
        self._dataframe_to_sheet(df, sheet)
    
    def _execute_vlookup(self, sheet: Worksheet, params: Dict[str, Any]):
        """跨表查找（支持外部文件）"""
        lookup_column = params.get("lookup_column")
        target_sheet_name = params.get("target_sheet")
        target_lookup_column = params.get("target_lookup_column")
        target_return_column = params.get("target_return_column")
        new_column_name = params.get("new_column_name", "查找结果")
        source_file = params.get("source_file")  # 外部文件路径（可选）
        
        # 检查是否是外部文件引用（格式：文件名!工作表名）
        external_wb = None
        if source_file:
            # 使用显式提供的外部文件路径
            source_path = Path(source_file)
            if not source_path.exists():
                raise ExecutionError(f"源文件不存在: {source_file}")
            external_wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
            # 如果 target_sheet_name 包含文件名前缀，去掉它
            if '!' in target_sheet_name:
                target_sheet_name = target_sheet_name.split('!')[-1]
            if target_sheet_name in external_wb.sheetnames:
                target_sheet = external_wb[target_sheet_name]
            else:
                target_sheet = external_wb.active
        elif '!' in target_sheet_name:
            # 格式：文件名!工作表名（但没有提供 source_file）
            raise ExecutionError(
                f"目标工作表格式错误: {target_sheet_name}。"
                f"多文件场景请使用 MERGE_HORIZONTAL 操作，或确保已提供 source_file 参数"
            )
        else:
            # 同一工作簿内的工作表
            if target_sheet_name not in self.workbook.sheetnames:
                raise ExecutionError(f"目标工作表不存在: {target_sheet_name}")
            target_sheet = self.workbook[target_sheet_name]
        
        # 构建查找表
        target_df = self._sheet_to_dataframe(target_sheet)
        lookup_dict = dict(zip(
            target_df[target_lookup_column].astype(str),
            target_df[target_return_column]
        ))
        
        # 在源表添加新列
        next_col = sheet.max_column + 1
        sheet.cell(row=1, column=next_col, value=new_column_name)
        
        lookup_col_idx = self._get_column_index(sheet, lookup_column)
        
        for row in range(2, sheet.max_row + 1):
            lookup_value = str(sheet.cell(row=row, column=lookup_col_idx).value)
            result = lookup_dict.get(lookup_value, "")
            sheet.cell(row=row, column=next_col, value=result)
        
        # 关闭外部工作簿
        if external_wb:
            external_wb.close()
    
    def _execute_pivot(self, sheet: Worksheet, params: Dict[str, Any]):
        """数据透视（创建新工作表）"""
        index_col = params.get("index")
        columns_col = params.get("columns")
        values_col = params.get("values")
        aggfunc = params.get("aggfunc", "sum")
        
        df = self._sheet_to_dataframe(sheet)
        
        # 验证列名是否存在
        available_cols = list(df.columns)
        
        # 辅助函数：验证单个列名
        def validate_column(col_name, col_label):
            if not col_name:
                return
            
            # 如果是列表，验证列表中的每一列
            if isinstance(col_name, list):
                for c in col_name:
                    if c not in available_cols:
                        matches = [ac for ac in available_cols if isinstance(ac, str) and isinstance(c, str) and (c.lower() in ac.lower() or ac.lower() in c.lower())]
                        if matches:
                            raise ExecutionError(
                                f"找不到{col_label} '{c}'，您可能是指: {matches[:3]}"
                            )
                        else:
                            raise ExecutionError(
                                f"找不到{col_label} '{c}'。可用的列: {available_cols[:10]}"
                            )
            else:
                # 单个列名验证
                if col_name not in available_cols:
                    matches = [c for c in available_cols if isinstance(c, str) and isinstance(col_name, str) and (col_name.lower() in c.lower() or c.lower() in col_name.lower())]
                    if matches:
                        raise ExecutionError(
                            f"找不到{col_label} '{col_name}'，您可能是指: {matches[:3]}"
                        )
                    else:
                        raise ExecutionError(
                            f"找不到{col_label} '{col_name}'。可用的列: {available_cols[:10]}"
                        )
        
        # 验证各个列
        validate_column(index_col, "行标签列")
        validate_column(columns_col, "列标签列")
        validate_column(values_col, "值列")
        
        # 创建数据透视表
        try:
            pivot_df = pd.pivot_table(
                df,
                index=index_col,
                columns=columns_col if columns_col else None,
                values=values_col,
                aggfunc=aggfunc
            ).reset_index()
        except Exception as e:
            raise ExecutionError(f"创建透视表失败: {str(e)}")
        
        # 创建新工作表存放透视结果
        pivot_sheet_name = f"{sheet.title}_透视表"
        if pivot_sheet_name in self.workbook.sheetnames:
            del self.workbook[pivot_sheet_name]
        
        pivot_sheet = self.workbook.create_sheet(pivot_sheet_name)
        
        # 写入表头
        for c_idx, col_name in enumerate(pivot_df.columns, start=1):
            pivot_sheet.cell(row=1, column=c_idx, value=str(col_name))
        
        # 写入数据
        for r_idx, row in enumerate(pivot_df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                pivot_sheet.cell(row=r_idx, column=c_idx, value=value)
        
        self._log(f"  透视表已创建: {pivot_sheet_name}，共 {len(pivot_df)} 行")
    
    def _execute_create_chart(self, sheet: Worksheet, params: Dict[str, Any]):
        """创建图表（嵌入Excel并生成图片）"""
        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm
        from io import BytesIO
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.chart import (
            LineChart, BarChart, PieChart, ScatterChart, 
            AreaChart, Reference
        )
        
        chart_type = params.get("chart_type", "bar").lower()
        data_columns = params.get("data_columns", [])
        label_column = params.get("label_column", "")
        title = params.get("title", "图表")
        position = params.get("position", "new_sheet")
        width = params.get("width", 15)
        height = params.get("height", 10)
        sheet_name = params.get("sheet_name", f"图表_{chart_type}")
        
        if not data_columns:
            raise ExecutionError("必须指定至少一个数据列")
        
        # 验证列是否存在
        headers = [cell.value for cell in sheet[1]]
        self._log(f"  图表数据 - 可用列: {[h for h in headers if h]}")
        self._log(f"  图表数据 - 请求的数据列: {data_columns}, 标签列: {label_column}")
        
        # 智能列名匹配函数
        def resolve_column_name(col_name, header_list):
            """尝试智能匹配列名，返回匹配到的实际列名或 None"""
            if not col_name:
                return None
            # 1. 精确匹配
            if col_name in header_list:
                return col_name
            # 2. 大小写不敏感匹配
            for h in header_list:
                if h and str(h).lower() == str(col_name).lower():
                    return h
            # 3. 包含匹配（列名包含关键字或反过来）
            for h in header_list:
                if h and isinstance(h, str):
                    if col_name.lower() in h.lower() or h.lower() in col_name.lower():
                        return h
            return None
        
        # 自动修正数据列名
        resolved_data_columns = []
        for col in data_columns:
            resolved = resolve_column_name(col, headers)
            if resolved:
                resolved_data_columns.append(resolved)
                if resolved != col:
                    self._log(f"  列名自动修正: '{col}' -> '{resolved}'")
            else:
                error_msg = f"找不到数据列: '{col}'"
                available = [h for h in headers if h][:10]
                error_msg += f"\n可用的列: {available}"
                raise ExecutionError(error_msg)
        
        data_columns = resolved_data_columns
        
        # 自动修正标签列名
        if label_column:
            resolved_label = resolve_column_name(label_column, headers)
            if resolved_label:
                if resolved_label != label_column:
                    self._log(f"  标签列名自动修正: '{label_column}' -> '{resolved_label}'")
                label_column = resolved_label
            else:
                error_msg = f"找不到标签列: '{label_column}'"
                available = [h for h in headers if h][:10]
                error_msg += f"\n可用的列: {available}"
                raise ExecutionError(error_msg)
        
        # 提取数据 - 使用公式计算后的值
        df = self._sheet_to_dataframe_with_values(sheet)
        self._log(f"  图表数据 - DataFrame 形状: {df.shape}, 行数: {len(df)}")
        
        # #region agent log
        import json
        with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
            f.write(json.dumps({"location":"excel_executor.py:1049","message":"chart_dataframe_loaded","data":{"shape":str(df.shape),"columns":list(df.columns)[:10],"row_count":len(df)},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"C"}) + '\n')
        # #endregion
        
        # 验证数据不为空
        if len(df) == 0:
            raise ExecutionError("数据表为空，无法创建图表")
        
        # 检查数据列是否有有效数据
        for col in data_columns:
            if col in df.columns:
                non_null = df[col].dropna()
                self._log(f"  列 '{col}' 数据: 非空值 {len(non_null)} 个, 示例: {non_null.head(3).tolist()}")
                
                # #region agent log
                with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
                    f.write(json.dumps({"location":"excel_executor.py:1065","message":"chart_column_data","data":{"column":col,"non_null_count":len(non_null),"total_count":len(df[col]),"sample_values":non_null.head(3).tolist()},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"C"}) + '\n')
                # #endregion
        
        # 处理中文字体（matplotlib 显示中文）
        try:
            plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'DejaVu Sans']
            plt.rcParams['axes.unicode_minus'] = False
        except Exception:
            pass
        
        # 创建matplotlib图表
        fig, ax = plt.subplots(figsize=(width, height))
        
        # 准备数据
        if label_column:
            labels = df[label_column].astype(str).tolist()
        else:
            labels = [f"行{i+1}" for i in range(len(df))]
        
        try:
            if chart_type == "pie":
                # 饼图：只使用第一个数据列
                values = pd.to_numeric(df[data_columns[0]], errors='coerce').fillna(0)
                ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
                ax.set_title(title, fontsize=16, fontweight='bold')
                
            elif chart_type == "line":
                # 折线图
                x = range(len(labels))
                for col in data_columns:
                    y = pd.to_numeric(df[col], errors='coerce').fillna(0)
                    ax.plot(x, y, marker='o', label=col, linewidth=2)
                ax.set_xticks(x)
                ax.set_xticklabels(labels, rotation=45, ha='right')
                ax.set_title(title, fontsize=16, fontweight='bold')
                ax.legend()
                ax.grid(True, alpha=0.3)
                
            elif chart_type in ["bar", "column"]:
                # 柱状图
                x = range(len(labels))
                bar_width = 0.8 / len(data_columns)
                bars_list = []
                for i, col in enumerate(data_columns):
                    y = pd.to_numeric(df[col], errors='coerce').fillna(0)
                    offset = (i - len(data_columns)/2 + 0.5) * bar_width
                    bars = ax.bar([pos + offset for pos in x], y, bar_width, label=col)
                    bars_list.append((bars, y))
                
                # 添加数据标签
                show_values = params.get("show_values", True)  # 默认显示数据标签
                
                # #region agent log
                import json
                with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
                    f.write(json.dumps({"location":"excel_executor.py:1110","message":"chart_show_values_param","data":{"show_values":show_values,"params_keys":list(params.keys())},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"D"}) + '\n')
                # #endregion
                
                if show_values:
                    for bars, y_values in bars_list:
                        for bar, val in zip(bars, y_values):
                            height = bar.get_height()
                            ax.annotate(f'{val:.0f}' if val == int(val) else f'{val:.1f}',
                                       xy=(bar.get_x() + bar.get_width() / 2, height),
                                       xytext=(0, 3),  # 3 点向上偏移
                                       textcoords="offset points",
                                       ha='center', va='bottom',
                                       fontsize=9, fontweight='bold')
                
                ax.set_xticks(x)
                ax.set_xticklabels(labels, rotation=45, ha='right')
                ax.set_title(title, fontsize=16, fontweight='bold')
                if len(data_columns) > 1:
                    ax.legend()
                ax.grid(True, alpha=0.3, axis='y')
                
            elif chart_type == "scatter":
                # 散点图：需要至少两列数据
                if len(data_columns) < 2:
                    raise ExecutionError("散点图需要至少两列数据")
                x_data = pd.to_numeric(df[data_columns[0]], errors='coerce').fillna(0)
                y_data = pd.to_numeric(df[data_columns[1]], errors='coerce').fillna(0)
                ax.scatter(x_data, y_data, alpha=0.6, s=100)
                ax.set_xlabel(data_columns[0], fontsize=12)
                ax.set_ylabel(data_columns[1], fontsize=12)
                ax.set_title(title, fontsize=16, fontweight='bold')
                ax.grid(True, alpha=0.3)
                
            elif chart_type == "area":
                # 面积图
                x = range(len(labels))
                for col in data_columns:
                    y = pd.to_numeric(df[col], errors='coerce').fillna(0)
                    ax.fill_between(x, y, alpha=0.5, label=col)
                ax.set_xticks(x)
                ax.set_xticklabels(labels, rotation=45, ha='right')
                ax.set_title(title, fontsize=16, fontweight='bold')
                ax.legend()
                ax.grid(True, alpha=0.3)
                
            else:
                raise ExecutionError(f"不支持的图表类型: {chart_type}")
            
            plt.tight_layout()
            
            # 保存为图片（用于嵌入Excel）
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close(fig)
            
            # 重置缓冲区位置并创建图像对象
            img_buffer.seek(0)
            img_data = img_buffer.read()
            img_buffer.close()
            
            # 使用新的缓冲区创建图像，确保数据独立
            img_stream = BytesIO(img_data)
            
            # #region agent log
            import json
            with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
                f.write(json.dumps({"location":"excel_executor.py:1171","message":"chart_position_check","data":{"position":position,"sheet_name":sheet_name},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"E"}) + '\n')
            # #endregion
            
            # 嵌入Excel
            if position == "new_sheet":
                # 创建新工作表
                if sheet_name in self.workbook.sheetnames:
                    # 如果已存在，删除旧的
                    del self.workbook[sheet_name]
                chart_sheet = self.workbook.create_sheet(sheet_name)
                img = OpenpyxlImage(img_stream)
                img.anchor = 'A1'
                chart_sheet.add_image(img)
                self._log(f"  图表已创建在新工作表: {sheet_name}")
            else:
                # 嵌入当前工作表
                img = OpenpyxlImage(img_stream)
                # 放置在数据表右侧
                img.anchor = f'{get_column_letter(sheet.max_column + 2)}1'
                sheet.add_image(img)
                self._log(f"  图表已嵌入当前工作表")
                
        except Exception as e:
            raise ExecutionError(f"创建图表失败: {str(e)}")
    
    def _execute_merge_vertical(self, sheet: Worksheet, params: Dict[str, Any]):
        """纵向合并：将外部文件的数据追加到当前表格下方"""
        source_file = params.get("source_file")  # 源文件路径
        source_sheet_name = params.get("source_sheet", None)  # 源工作表名
        skip_header = params.get("skip_header", True)  # 是否跳过源文件表头
        
        # #region agent log
        import json
        with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
            f.write(json.dumps({"location":"excel_executor.py:1195","message":"merge_vertical_start","data":{"source_file":source_file,"skip_header":skip_header},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"A"}) + '\n')
        # #endregion
        
        if not source_file:
            raise ExecutionError("纵向合并需要指定源文件路径 (source_file)")
        
        source_path = Path(source_file)
        if not source_path.exists():
            raise ExecutionError(f"源文件不存在: {source_file}")
        
        try:
            # 加载源文件
            source_wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
            
            # 获取源工作表
            if source_sheet_name and source_sheet_name in source_wb.sheetnames:
                source_sheet = source_wb[source_sheet_name]
            else:
                source_sheet = source_wb.active
            
            # 获取当前表的最后一行和表头
            current_max_row = sheet.max_row
            target_headers = [str(cell.value).lower().strip() if cell.value else '' for cell in sheet[1]]
            
            # #region agent log
            with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
                f.write(json.dumps({"location":"excel_executor.py:1220","message":"headers_detected","data":{"target_headers":target_headers[:5],"current_max_row":current_max_row},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"A"}) + '\n')
            # #endregion
            
            # 获取源表数据
            rows_added = 0
            start_row = 2 if skip_header else 1  # 跳过源文件表头
            
            for row_idx, row in enumerate(source_sheet.iter_rows(min_row=start_row), start=1):
                # 获取行数据
                row_values = [cell.value for cell in row]
                
                # 智能检测：如果这行看起来像表头（与目标表头匹配），则跳过
                row_as_headers = [str(v).lower().strip() if v else '' for v in row_values]
                
                # #region agent log
                if row_idx <= 3:  # 只记录前3行
                    with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
                        f.write(json.dumps({"location":"excel_executor.py:1235","message":"row_comparison","data":{"row_idx":row_idx,"row_as_headers":row_as_headers[:5],"matches":row_as_headers == target_headers[:len(row_as_headers)]},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"A"}) + '\n')
                # #endregion
                
                if row_as_headers == target_headers[:len(row_as_headers)]:
                    self._log(f"  检测到重复表头，自动跳过")
                    continue
                
                # 追加数据
                for col_idx, cell in enumerate(row, start=1):
                    target_row = current_max_row + rows_added + 1
                    sheet.cell(row=target_row, column=col_idx, value=cell.value)
                rows_added += 1
            
            # #region agent log
            with open('/Users/louis/PycharmProjects/Open Source/LLM-Excel-Copilot/.cursor/debug.log', 'a') as f:
                f.write(json.dumps({"location":"excel_executor.py:1252","message":"merge_vertical_complete","data":{"rows_added":rows_added},"timestamp":datetime.now().timestamp()*1000,"sessionId":"debug-session","hypothesisId":"A"}) + '\n')
            # #endregion
            
            source_wb.close()
            self._log(f"  纵向合并完成：从 {source_path.name} 追加了 {rows_added} 行数据")
            
        except Exception as e:
            raise ExecutionError(f"纵向合并失败: {str(e)}")
    
    def _execute_merge_horizontal(self, sheet: Worksheet, params: Dict[str, Any]):
        """横向合并：按关键列匹配，将外部文件的列添加到当前表格"""
        source_file = params.get("source_file")  # 源文件路径
        source_sheet_name = params.get("source_sheet", None)
        key_column = params.get("key_column")  # 当前表的关键列
        source_key_column = params.get("source_key_column")  # 源表的关键列
        columns_to_add = params.get("columns_to_add", [])  # 要添加的列名列表
        
        if not source_file:
            raise ExecutionError("横向合并需要指定源文件路径 (source_file)")
        if not key_column:
            raise ExecutionError("横向合并需要指定当前表的关键列 (key_column)")
        
        source_path = Path(source_file)
        if not source_path.exists():
            raise ExecutionError(f"源文件不存在: {source_file}")
        
        try:
            # 获取当前表的关键列索引
            key_col_idx = self._get_column_index(sheet, key_column)
            
            # 读取源文件
            source_wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
            if source_sheet_name and source_sheet_name in source_wb.sheetnames:
                source_sheet = source_wb[source_sheet_name]
            else:
                source_sheet = source_wb.active
            
            # 将源表转换为 DataFrame
            source_data = list(source_sheet.values)
            if not source_data:
                raise ExecutionError("源文件没有数据")
            
            source_headers = [str(h) if h else f"列{i}" for i, h in enumerate(source_data[0])]
            source_df = pd.DataFrame(source_data[1:], columns=source_headers)
            
            # 确定源表关键列
            src_key = source_key_column or key_column
            if src_key not in source_df.columns:
                raise ExecutionError(f"源表中找不到关键列: {src_key}")
            
            # 构建查找字典
            # 如果没有指定 columns_to_add，则添加所有非关键列
            if not columns_to_add:
                columns_to_add = [c for c in source_df.columns if c != src_key]
            
            lookup_dict = {}
            for _, row in source_df.iterrows():
                key_val = str(row[src_key])
                lookup_dict[key_val] = {col: row[col] for col in columns_to_add if col in source_df.columns}
            
            # 在当前表添加新列（处理重复列名）
            current_max_col = sheet.max_column
            current_headers = [sheet.cell(row=1, column=c).value for c in range(1, current_max_col + 1)]
            added_cols = []
            col_name_mapping = {}  # 原始列名 -> 实际使用的列名
            
            for i, col_name in enumerate(columns_to_add):
                if col_name in source_df.columns:
                    new_col_idx = current_max_col + i + 1
                    # 如果目标表已有同名列，添加后缀
                    actual_col_name = col_name
                    if col_name in current_headers:
                        suffix = 1
                        while f"{col_name}_{suffix}" in current_headers or f"{col_name}_{suffix}" in [c[0] for c in added_cols]:
                            suffix += 1
                        actual_col_name = f"{col_name}_{suffix}"
                        self._log(f"  列名冲突: '{col_name}' -> '{actual_col_name}'")
                    
                    sheet.cell(row=1, column=new_col_idx, value=actual_col_name)
                    added_cols.append((actual_col_name, new_col_idx, col_name))  # (实际列名, 列索引, 源列名)
                    col_name_mapping[col_name] = actual_col_name
            
            # 填充数据
            matches = 0
            for row_idx in range(2, sheet.max_row + 1):
                key_val = str(sheet.cell(row=row_idx, column=key_col_idx).value)
                if key_val in lookup_dict:
                    for actual_name, col_idx, source_name in added_cols:
                        value = lookup_dict[key_val].get(source_name)
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                    matches += 1
            
            source_wb.close()
            # 输出新列名以便后续操作使用
            new_col_names = [c[0] for c in added_cols]
            self._log(f"  横向合并完成：添加了 {len(added_cols)} 列 ({', '.join(new_col_names)})，匹配了 {matches} 行")
            
        except Exception as e:
            if isinstance(e, ExecutionError):
                raise
            raise ExecutionError(f"横向合并失败: {str(e)}")

    def get_log(self) -> List[str]:
        """获取操作日志"""
        return self.operation_log
    
    def get_operation_history(self) -> List[Dict[str, Any]]:
        """获取操作历史记录"""
        return self.operation_history
    
    def get_backup_path(self) -> Optional[Path]:
        """获取备份文件路径"""
        return self._backup_file if self._backup_file and self._backup_file.exists() else None
    
    def close(self):
        """关闭工作簿并清理临时文件"""
        self.workbook.close()
        
        # 清理临时文件
        if self._temp_file and self._temp_file.exists():
            try:
                self._temp_file.unlink()
            except Exception:
                pass
