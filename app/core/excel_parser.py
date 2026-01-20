"""
Excel 解析器模块
负责读取 Excel 文件并提取结构化元数据（不含具体隐私数据）
"""

import os
import re
from pathlib import Path
from typing import List, Optional, Tuple, Any
from datetime import datetime, date

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

# 尝试导入 xlrd 以支持 .xls 格式
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

from app.models import ExcelMetadata, SheetInfo, ColumnInfo


class ExcelParser:
    """
    Excel 文件解析器
    
    核心功能：
    1. 读取 Excel 文件（支持 .xlsx 和 .xls）
    2. 提取文件结构信息（工作表、表头、列类型等）
    3. 生成隐私安全的元数据描述（不包含具体数据值）
    """
    
    # 数据类型推断的采样行数
    SAMPLE_ROWS = 10
    
    def __init__(self, file_path: str | Path):
        """
        初始化解析器
        
        Args:
            file_path: Excel 文件路径
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        self.file_name = self.file_path.name
        self.file_size = self.file_path.stat().st_size
        self.extension = self.file_path.suffix.lower()
        
        # 验证文件格式
        if self.extension not in ['.xlsx', '.xls']:
            raise ValueError(f"不支持的文件格式: {self.extension}")
        
        if self.extension == '.xls' and not XLRD_AVAILABLE:
            raise RuntimeError("需要安装 xlrd 库来支持 .xls 格式")
    
    def parse(self, file_id: str) -> ExcelMetadata:
        """
        解析 Excel 文件，提取元数据
        
        Args:
            file_id: 文件唯一标识
            
        Returns:
            ExcelMetadata: 文件元数据
        """
        if self.extension == '.xlsx':
            return self._parse_xlsx(file_id)
        else:
            return self._parse_xls(file_id)
    
    def _parse_xlsx(self, file_id: str) -> ExcelMetadata:
        """解析 .xlsx 文件"""
        # 不使用 read_only 模式，以便访问合并单元格等信息
        workbook = openpyxl.load_workbook(self.file_path, read_only=False, data_only=False)
        
        sheets_info = []
        for idx, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            sheet_info = self._analyze_sheet_xlsx(sheet, idx)
            sheets_info.append(sheet_info)
        
        active_sheet = workbook.active.title if workbook.active else workbook.sheetnames[0]
        workbook.close()
        
        return ExcelMetadata(
            file_id=file_id,
            file_name=self.file_name,
            file_size=self.file_size,
            sheets=sheets_info,
            active_sheet=active_sheet
        )
    
    def _analyze_sheet_xlsx(self, sheet: Worksheet, index: int) -> SheetInfo:
        """分析 xlsx 工作表"""
        # 获取实际数据范围
        rows = list(sheet.iter_rows())
        if not rows:
            return SheetInfo(
                name=sheet.title,
                index=index,
                total_rows=0,
                total_cols=0,
                headers=[],
                columns=[],
                has_merged_cells=bool(sheet.merged_cells.ranges),
                has_formulas=False
            )
        
        # 提取表头（第一行）
        header_row = rows[0]
        headers = [self._get_cell_value_safe(cell) for cell in header_row]
        # 规范化列名：去除换行符，替换为空字符（Excel 中常见的多行表头）
        headers = [str(h).replace('\n', '').replace('\r', '') if h else f"列{i+1}" for i, h in enumerate(headers)]
        
        # 去除末尾空列
        while headers and headers[-1].startswith("列"):
            if all(self._get_cell_value_safe(rows[r][len(headers)-1]) == "" 
                   for r in range(min(5, len(rows))) if len(rows[r]) > len(headers)-1):
                headers.pop()
            else:
                break
        
        total_cols = len(headers)
        total_rows = len(rows) - 1  # 减去表头行
        
        # 分析每列
        columns_info = []
        has_formulas = False
        
        for col_idx, header in enumerate(headers):
            # 采样数据用于类型推断
            sample_values = []
            for row_idx in range(1, min(self.SAMPLE_ROWS + 1, len(rows))):
                if col_idx < len(rows[row_idx]):
                    cell = rows[row_idx][col_idx]
                    value = self._get_cell_value_safe(cell)
                    sample_values.append(value)
                    # 检测公式
                    if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                        has_formulas = True
            
            # 推断数据类型
            data_type = self._infer_data_type(sample_values)
            
            # 统计空值和唯一值
            has_empty = any(v == "" or v is None for v in sample_values)
            unique_count = len(set(v for v in sample_values if v))
            
            columns_info.append(ColumnInfo(
                name=header,
                index=col_idx,
                data_type=data_type,
                sample_values=self._mask_sample_values(sample_values, data_type),
                has_empty=has_empty,
                unique_count=unique_count if unique_count < 10 else None
            ))
        
        return SheetInfo(
            name=sheet.title,
            index=index,
            total_rows=total_rows,
            total_cols=total_cols,
            headers=headers,
            columns=columns_info,
            has_merged_cells=bool(sheet.merged_cells.ranges),
            has_formulas=has_formulas
        )
    
    def _parse_xls(self, file_id: str) -> ExcelMetadata:
        """解析 .xls 文件"""
        workbook = xlrd.open_workbook(self.file_path)
        
        sheets_info = []
        for idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(idx)
            sheet_info = self._analyze_sheet_xls(sheet, idx)
            sheets_info.append(sheet_info)
        
        return ExcelMetadata(
            file_id=file_id,
            file_name=self.file_name,
            file_size=self.file_size,
            sheets=sheets_info,
            active_sheet=workbook.sheet_by_index(0).name
        )
    
    def _analyze_sheet_xls(self, sheet, index: int) -> SheetInfo:
        """分析 xls 工作表"""
        if sheet.nrows == 0:
            return SheetInfo(
                name=sheet.name,
                index=index,
                total_rows=0,
                total_cols=0,
                headers=[],
                columns=[],
                has_merged_cells=False,
                has_formulas=False
            )
        
        # 提取表头
        headers = [str(sheet.cell_value(0, c)) or f"列{c+1}" for c in range(sheet.ncols)]
        # 规范化列名：去除换行符（Excel 中常见的多行表头）
        headers = [h.replace('\n', '').replace('\r', '') for h in headers]
        total_cols = len(headers)
        total_rows = sheet.nrows - 1
        
        # 分析每列
        columns_info = []
        for col_idx, header in enumerate(headers):
            sample_values = []
            for row_idx in range(1, min(self.SAMPLE_ROWS + 1, sheet.nrows)):
                value = sheet.cell_value(row_idx, col_idx)
                sample_values.append(str(value) if value else "")
            
            data_type = self._infer_data_type(sample_values)
            has_empty = any(v == "" for v in sample_values)
            unique_count = len(set(v for v in sample_values if v))
            
            columns_info.append(ColumnInfo(
                name=header,
                index=col_idx,
                data_type=data_type,
                sample_values=self._mask_sample_values(sample_values, data_type),
                has_empty=has_empty,
                unique_count=unique_count if unique_count < 10 else None
            ))
        
        return SheetInfo(
            name=sheet.name,
            index=index,
            total_rows=total_rows,
            total_cols=total_cols,
            headers=headers,
            columns=columns_info,
            has_merged_cells=bool(sheet.merged_cells),
            has_formulas=False  # xlrd 不直接暴露公式信息
        )
    
    def _get_cell_value_safe(self, cell: Cell) -> str:
        """安全获取单元格值"""
        if cell.value is None:
            return ""
        if isinstance(cell.value, (datetime, date)):
            return cell.value.strftime("%Y-%m-%d")
        return str(cell.value).strip()
    
    def _infer_data_type(self, values: List[str]) -> str:
        """
        根据采样值推断列的数据类型
        
        Returns:
            str: 类型名称 (数字/文本/日期/布尔/混合)
        """
        if not values or all(v == "" for v in values):
            return "空"
        
        non_empty = [v for v in values if v]
        if not non_empty:
            return "空"
        
        type_counts = {"数字": 0, "日期": 0, "布尔": 0, "文本": 0}
        
        for value in non_empty:
            if self._is_number(value):
                type_counts["数字"] += 1
            elif self._is_date(value):
                type_counts["日期"] += 1
            elif value.lower() in ("true", "false", "是", "否", "1", "0"):
                type_counts["布尔"] += 1
            else:
                type_counts["文本"] += 1
        
        # 如果某种类型占比超过 80%，则认定为该类型
        total = len(non_empty)
        for dtype, count in type_counts.items():
            if count / total >= 0.8:
                return dtype
        
        return "混合"
    
    def _is_number(self, value: str) -> bool:
        """判断是否为数字"""
        try:
            # 处理千分位逗号
            cleaned = value.replace(",", "").replace("¥", "").replace("$", "").strip()
            float(cleaned)
            return True
        except ValueError:
            return False
    
    def _is_date(self, value: str) -> bool:
        """判断是否为日期"""
        date_patterns = [
            r"\d{4}[-/]\d{1,2}[-/]\d{1,2}",  # 2024-01-01
            r"\d{1,2}[-/]\d{1,2}[-/]\d{4}",  # 01/01/2024
            r"\d{4}年\d{1,2}月\d{1,2}日",     # 2024年1月1日
        ]
        for pattern in date_patterns:
            if re.match(pattern, value):
                return True
        return False
    
    def _mask_sample_values(self, values: List[str], data_type: str) -> List[str]:
        """
        对采样值进行智能脱敏处理
        
        隐私保护策略：
        - 对于枚举类型（选项较少的分类列），显示实际值（通常不含隐私）
        - 对于长文本（可能是姓名、地址等），只显示格式
        """
        if not values or all(v == "" for v in values):
            return []
        
        # 过滤空值
        non_empty_values = [v for v in values if v]
        if not non_empty_values:
            return []
        
        # 统计唯一值数量，判断是否为枚举类型
        unique_values = list(set(non_empty_values[:20]))  # 最多检查前20个值
        unique_count = len(unique_values)
        total_count = len(non_empty_values[:20])
        
        # 如果唯一值较少（占比 < 50%），且文本较短，可能是枚举类型
        is_enum_like = (unique_count <= 8 and 
                       unique_count / max(total_count, 1) < 0.5 and
                       all(len(str(v)) <= 30 for v in unique_values[:5]))
        
        masked = []
        
        if data_type == "数字":
            # 只展示数值范围特征
            try:
                nums = [float(str(v).replace(",", "")) for v in non_empty_values[:3]]
                for num in nums:
                    if num == int(num):
                        masked.append(f"整数({len(str(int(num)))}位)")
                    else:
                        masked.append(f"小数")
            except Exception:
                masked.append("数字格式")
        
        elif data_type == "日期":
            masked.append("日期格式")
        
        elif data_type == "文本":
            if is_enum_like:
                # 枚举类型：显示实际值（最多5个）
                for v in unique_values[:5]:
                    if len(str(v)) <= 30:  # 确保不是太长的文本
                        masked.append(f'"{v}"')
                if unique_count > 5:
                    masked.append(f"... (共{unique_count}种)")
            else:
                # 非枚举类型：只显示长度特征（可能含隐私）
                for v in non_empty_values[:3]:
                    length = len(v)
                    if length <= 5:
                        masked.append("短文本(1-5字)")
                    elif length <= 20:
                        masked.append("中等文本(6-20字)")
                    else:
                        masked.append(f"长文本({length}字)")
        else:
            masked.append(data_type)
        
        return masked[:6]  # 最多返回6个示例
    
    def generate_description(self, metadata: ExcelMetadata) -> str:
        """
        生成供 LLM 理解的文件描述
        
        这个描述将发送给 LLM，帮助其理解文件结构，但不包含任何具体数据
        """
        lines = [
            f"## Excel 文件结构分析",
            f"",
            f"**文件名**: {metadata.file_name}",
            f"**工作表数量**: {len(metadata.sheets)}",
            f"",
        ]
        
        for sheet in metadata.sheets:
            lines.append(f"### 工作表: {sheet.name}")
            lines.append(f"- 数据行数: {sheet.total_rows} 行")
            lines.append(f"- 列数: {sheet.total_cols} 列")
            if sheet.has_merged_cells:
                lines.append(f"- ⚠️ 包含合并单元格")
            if sheet.has_formulas:
                lines.append(f"- 📐 包含公式")
            lines.append("")
            lines.append("**列信息**:")
            lines.append("")
            lines.append("| 序号 | 列名 | 数据类型 | 示例值 | 有空值 |")
            lines.append("|------|------|----------|--------|--------|")
            
            for col in sheet.columns:
                empty_mark = "✓" if col.has_empty else ""
                # 获取脱敏后的示例值
                sample_display = ", ".join(col.sample_values[:4]) if col.sample_values else "-"
                if len(sample_display) > 50:
                    sample_display = sample_display[:47] + "..."
                lines.append(f"| {col.index + 1} | {col.name} | {col.data_type} | {sample_display} | {empty_mark} |")
            
            lines.append("")
        
        return "\n".join(lines)
