"""
Excel Parser 单元测试
测试 Excel 文件结构解析功能
"""

import os
import pytest
import openpyxl
from pathlib import Path
from tempfile import NamedTemporaryFile

from app.core.excel_parser import ExcelParser
from app.models import ExcelMetadata, SheetInfo, ColumnInfo


class TestExcelParser:
    """ExcelParser 测试用例"""

    @pytest.fixture
    def sample_xlsx(self, tmp_path):
        """创建测试用的 xlsx 文件"""
        file_path = tmp_path / "test_sample.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "测试表"
        
        # 添加表头
        headers = ["姓名", "年龄", "部门", "薪资"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # 添加数据
        data = [
            ["张三", 28, "技术部", 15000],
            ["李四", 32, "市场部", 12000],
            ["王五", 25, "技术部", 10000],
            ["赵六", 35, "财务部", 18000],
        ]
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        workbook.save(file_path)
        return file_path

    @pytest.fixture
    def empty_xlsx(self, tmp_path):
        """创建空的 xlsx 文件"""
        file_path = tmp_path / "empty.xlsx"
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        return file_path

    def test_parse_basic_xlsx(self, sample_xlsx):
        """测试基本的 xlsx 文件解析"""
        parser = ExcelParser(sample_xlsx)
        metadata = parser.parse("test-file-id")
        
        assert metadata.file_id == "test-file-id"
        assert metadata.file_name == "test_sample.xlsx"
        assert len(metadata.sheets) == 1
        
        sheet = metadata.sheets[0]
        assert sheet.name == "测试表"
        assert sheet.total_rows == 4  # 不含表头
        assert sheet.total_cols == 4
        assert sheet.headers == ["姓名", "年龄", "部门", "薪资"]

    def test_parse_column_types(self, sample_xlsx):
        """测试列类型推断"""
        parser = ExcelParser(sample_xlsx)
        metadata = parser.parse("test-file-id")
        
        columns = metadata.sheets[0].columns
        
        # 姓名应该是文本类型
        assert columns[0].name == "姓名"
        assert columns[0].data_type == "文本"
        
        # 年龄应该是数字类型
        assert columns[1].name == "年龄"
        assert columns[1].data_type == "数字"
        
        # 薪资应该是数字类型
        assert columns[3].name == "薪资"
        assert columns[3].data_type == "数字"

    def test_parse_empty_sheet(self, empty_xlsx):
        """测试空工作表解析"""
        parser = ExcelParser(empty_xlsx)
        metadata = parser.parse("empty-file")
        
        assert len(metadata.sheets) == 1
        sheet = metadata.sheets[0]
        assert sheet.total_rows == 0
        assert sheet.total_cols == 0

    def test_file_not_found(self, tmp_path):
        """测试文件不存在的错误处理"""
        with pytest.raises(FileNotFoundError):
            ExcelParser(tmp_path / "nonexistent.xlsx")

    def test_invalid_format(self, tmp_path):
        """测试无效文件格式"""
        invalid_file = tmp_path / "invalid.txt"
        invalid_file.write_text("This is not an Excel file")
        
        with pytest.raises(ValueError, match="不支持的文件格式"):
            ExcelParser(invalid_file)

    def test_generate_description(self, sample_xlsx):
        """测试文件描述生成"""
        parser = ExcelParser(sample_xlsx)
        metadata = parser.parse("test-file-id")
        description = parser.generate_description(metadata)
        
        assert "test_sample.xlsx" in description
        assert "测试表" in description
        assert "4 行" in description  # 数据行数
        assert "姓名" in description
        assert "年龄" in description

    def test_multisheet_xlsx(self, tmp_path):
        """测试多工作表文件"""
        file_path = tmp_path / "multisheet.xlsx"
        workbook = openpyxl.Workbook()
        
        # 第一个工作表
        sheet1 = workbook.active
        sheet1.title = "销售数据"
        sheet1["A1"] = "产品"
        sheet1["B1"] = "销量"
        sheet1["A2"] = "产品A"
        sheet1["B2"] = 100
        
        # 第二个工作表
        sheet2 = workbook.create_sheet("库存数据")
        sheet2["A1"] = "产品"
        sheet2["B1"] = "库存"
        sheet2["A2"] = "产品A"
        sheet2["B2"] = 50
        
        workbook.save(file_path)
        
        parser = ExcelParser(file_path)
        metadata = parser.parse("multi-sheet-id")
        
        assert len(metadata.sheets) == 2
        assert metadata.sheets[0].name == "销售数据"
        assert metadata.sheets[1].name == "库存数据"

    def test_merged_cells_detection(self, tmp_path):
        """测试合并单元格检测"""
        file_path = tmp_path / "merged.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        sheet["A1"] = "标题"
        sheet.merge_cells("A1:C1")
        sheet["A2"] = "数据1"
        sheet["B2"] = "数据2"
        sheet["C2"] = "数据3"
        
        workbook.save(file_path)
        
        parser = ExcelParser(file_path)
        metadata = parser.parse("merged-file")
        
        assert metadata.sheets[0].has_merged_cells is True

    def test_formula_detection(self, tmp_path):
        """测试公式检测"""
        file_path = tmp_path / "formula.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        sheet["A1"] = "数值"
        sheet["B1"] = "公式"
        sheet["A2"] = 10
        sheet["B2"] = "=A2*2"
        
        workbook.save(file_path)
        
        parser = ExcelParser(file_path)
        metadata = parser.parse("formula-file")
        
        assert metadata.sheets[0].has_formulas is True
