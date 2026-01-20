"""
Excel Executor 单元测试
测试 Excel 操作执行功能
"""

import os
import pytest
import openpyxl
from pathlib import Path
from tempfile import TemporaryDirectory

from app.core.excel_executor import ExcelExecutor, ExecutionError
from app.models import Operation, OperationPlan, OperationType


class TestExcelExecutor:
    """ExcelExecutor 测试用例"""

    @pytest.fixture
    def sample_xlsx(self, tmp_path):
        """创建测试用的 xlsx 文件"""
        file_path = tmp_path / "test_executor.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "数据"
        
        # 添加表头
        headers = ["姓名", "年龄", "部门", "薪资", "状态"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # 添加数据
        data = [
            ["张三", 28, "技术部", 15000, "在职"],
            ["李四", 32, "市场部", 12000, "在职"],
            ["王五", 25, "技术部", 10000, "离职"],
            ["赵六", 35, "财务部", 18000, "在职"],
            ["张三", 28, "技术部", 15000, "在职"],  # 重复行
        ]
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        workbook.save(file_path)
        return file_path

    @pytest.fixture
    def output_dir(self, tmp_path):
        """创建输出目录"""
        output = tmp_path / "output"
        output.mkdir()
        return output

    def test_filter_operation(self, sample_xlsx, output_dir):
        """测试筛选操作"""
        executor = ExcelExecutor(sample_xlsx)
        
        plan = OperationPlan(
            operations=[
                Operation(
                    type=OperationType.FILTER,
                    params={"column": "部门", "operator": "eq", "value": "技术部"},
                    description="筛选技术部员工"
                )
            ],
            summary="筛选技术部"
        )
        
        output_path = output_dir / "filtered.xlsx"
        executor.execute_plan(plan, output_path)
        executor.close()
        
        # 验证结果
        result = openpyxl.load_workbook(output_path)
        sheet = result.active
        
        # 应该只有3行数据（表头 + 3条技术部记录）
        assert sheet.max_row == 4  # 1 header + 3 data rows
        
        # 验证所有数据都是技术部
        for row in range(2, sheet.max_row + 1):
            assert sheet.cell(row=row, column=3).value == "技术部"
        
        result.close()

    def test_sort_operation(self, sample_xlsx, output_dir):
        """测试排序操作"""
        executor = ExcelExecutor(sample_xlsx)
        
        plan = OperationPlan(
            operations=[
                Operation(
                    type=OperationType.SORT,
                    params={"column": "薪资", "order": "desc"},
                    description="按薪资降序排序"
                )
            ],
            summary="排序"
        )
        
        output_path = output_dir / "sorted.xlsx"
        executor.execute_plan(plan, output_path)
        executor.close()
        
        # 验证结果
        result = openpyxl.load_workbook(output_path)
        sheet = result.active
        
        # 验证薪资是降序排列
        salaries = [sheet.cell(row=row, column=4).value for row in range(2, sheet.max_row + 1)]
        assert salaries == sorted(salaries, reverse=True)
        
        result.close()

    def test_deduplicate_operation(self, sample_xlsx, output_dir):
        """测试去重操作"""
        executor = ExcelExecutor(sample_xlsx)
        
        plan = OperationPlan(
            operations=[
                Operation(
                    type=OperationType.DEDUPLICATE,
                    params={"columns": ["姓名", "年龄"], "keep": "first"},
                    description="按姓名和年龄去重"
                )
            ],
            summary="去重"
        )
        
        output_path = output_dir / "dedup.xlsx"
        executor.execute_plan(plan, output_path)
        executor.close()
        
        # 验证结果
        result = openpyxl.load_workbook(output_path)
        sheet = result.active
        
        # 原来5行数据，有1行重复，去重后应该是4行 + 1表头
        assert sheet.max_row == 5
        
        result.close()

    def test_add_column_operation(self, sample_xlsx, output_dir):
        """测试新增列操作"""
        executor = ExcelExecutor(sample_xlsx)
        
        plan = OperationPlan(
            operations=[
                Operation(
                    type=OperationType.ADD_COLUMN,
                    params={"name": "备注", "position": "end"},
                    description="添加备注列"
                )
            ],
            summary="新增列"
        )
        
        output_path = output_dir / "added_col.xlsx"
        executor.execute_plan(plan, output_path)
        executor.close()
        
        # 验证结果
        result = openpyxl.load_workbook(output_path)
        sheet = result.active
        
        # 验证新列存在
        assert sheet.cell(row=1, column=6).value == "备注"
        
        result.close()

    def test_delete_rows_operation(self, sample_xlsx, output_dir):
        """测试删除行操作"""
        executor = ExcelExecutor(sample_xlsx)
        
        plan = OperationPlan(
            operations=[
                Operation(
                    type=OperationType.DELETE_ROWS,
                    params={
                        "condition": {
                            "column": "状态",
                            "operator": "eq",
                            "value": "离职"
                        }
                    },
                    description="删除离职员工"
                )
            ],
            summary="删除行"
        )
        
        output_path = output_dir / "deleted_rows.xlsx"
        executor.execute_plan(plan, output_path)
        executor.close()
        
        # 验证结果
        result = openpyxl.load_workbook(output_path)
        sheet = result.active
        
        # 原来5行数据，删除1行离职员工，应该剩4行 + 1表头
        assert sheet.max_row == 5
        
        # 验证没有离职员工
        for row in range(2, sheet.max_row + 1):
            assert sheet.cell(row=row, column=5).value != "离职"
        
        result.close()

    def test_replace_operation(self, sample_xlsx, output_dir):
        """测试替换操作"""
        executor = ExcelExecutor(sample_xlsx)
        
        plan = OperationPlan(
            operations=[
                Operation(
                    type=OperationType.REPLACE,
                    params={
                        "column": "部门",
                        "old_value": "技术部",
                        "new_value": "研发部"
                    },
                    description="将技术部改为研发部"
                )
            ],
            summary="替换"
        )
        
        output_path = output_dir / "replaced.xlsx"
        executor.execute_plan(plan, output_path)
        executor.close()
        
        # 验证结果
        result = openpyxl.load_workbook(output_path)
        sheet = result.active
        
        # 验证没有技术部，只有研发部
        departments = [sheet.cell(row=row, column=3).value for row in range(2, sheet.max_row + 1)]
        assert "技术部" not in departments
        assert "研发部" in departments
        
        result.close()

    def test_invalid_column_error(self, sample_xlsx, output_dir):
        """测试无效列名错误处理"""
        executor = ExcelExecutor(sample_xlsx)
        
        plan = OperationPlan(
            operations=[
                Operation(
                    type=OperationType.FILTER,
                    params={"column": "不存在的列", "operator": "eq", "value": "test"},
                    description="筛选不存在的列"
                )
            ],
            summary="错误测试"
        )
        
        with pytest.raises(ExecutionError, match="找不到列"):
            executor.execute_plan(plan, output_dir / "error.xlsx")
        
        executor.close()

    def test_file_not_found_error(self, tmp_path):
        """测试文件不存在错误"""
        with pytest.raises(FileNotFoundError):
            ExcelExecutor(tmp_path / "nonexistent.xlsx")

    def test_chained_operations(self, sample_xlsx, output_dir):
        """测试链式操作"""
        executor = ExcelExecutor(sample_xlsx)
        
        plan = OperationPlan(
            operations=[
                Operation(
                    type=OperationType.FILTER,
                    params={"column": "状态", "operator": "eq", "value": "在职"},
                    description="筛选在职员工"
                ),
                Operation(
                    type=OperationType.SORT,
                    params={"column": "薪资", "order": "desc"},
                    description="按薪资排序"
                ),
                Operation(
                    type=OperationType.DEDUPLICATE,
                    params={"columns": ["姓名"], "keep": "first"},
                    description="按姓名去重"
                )
            ],
            summary="链式操作测试"
        )
        
        output_path = output_dir / "chained.xlsx"
        executor.execute_plan(plan, output_path)
        executor.close()
        
        # 验证结果
        result = openpyxl.load_workbook(output_path)
        sheet = result.active
        
        # 应该有3个不重复的在职员工
        assert sheet.max_row == 4  # 1 header + 3 unique 在职 employees
        
        # 验证排序正确
        salaries = [sheet.cell(row=row, column=4).value for row in range(2, sheet.max_row + 1)]
        assert salaries == sorted(salaries, reverse=True)
        
        result.close()

    def test_calculate_operation(self, sample_xlsx, output_dir):
        """测试汇总计算操作"""
        executor = ExcelExecutor(sample_xlsx)
        
        plan = OperationPlan(
            operations=[
                Operation(
                    type=OperationType.CALCULATE,
                    params={
                        "operations": [
                            {"column": "薪资", "function": "sum"},
                            {"column": "薪资", "function": "avg"}
                        ]
                    },
                    description="计算薪资汇总"
                )
            ],
            summary="汇总计算"
        )
        
        output_path = output_dir / "calculated.xlsx"
        executor.execute_plan(plan, output_path)
        executor.close()
        
        # 验证结果
        result = openpyxl.load_workbook(output_path)
        sheet = result.active
        
        # 最后一行应该是汇总行
        assert sheet.cell(row=sheet.max_row, column=1).value == "汇总"
        
        result.close()
